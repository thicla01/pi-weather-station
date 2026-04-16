#!/usr/bin/env python3
"""
Pi Weather Station — Debug CSV to Excel converter
Splits each === SECTION === from the debug CSV export into a separate sheet.

Usage:
    python3 debug_csv_to_excel.py [input.csv] [output.xlsx]

    If no arguments are given, the most recent weather-station-debug-*.csv
    file in ~/Downloads is used, and the output is placed alongside it.

Requirements:
    pip3 install openpyxl
"""

import sys
import csv
import os
import glob
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Run:   pip3 install openpyxl")
    sys.exit(1)


# ---------- Styling ----------

SECTION_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FILL    = PatternFill("solid", fgColor="2E75B6")   # medium blue
SECTION_FONT   = Font(bold=True, color="FFFFFF", size=11)
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=10)
META_LABEL_FONT = Font(bold=True, size=10)


def style_section_row(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT


def style_header_row(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# ---------- Parsing ----------

def is_section_marker(row):
    return row and row[0].startswith("===") and row[0].endswith("===")


def section_name(row):
    return row[0].strip("= ").strip()


def parse_csv(path):
    """
    Returns a list of sections:
      [
        { "name": "Overview", "rows": [[label, value], ...] },
        { "name": "SERVER KPIs", "rows": [...] },
        ...
      ]
    The lines before the first === marker go into an "Overview" section.
    """
    sections = []
    current = {"name": "Overview", "rows": []}

    with open(path, newline="", encoding="utf-8-sig") as f:
        # Skip optional sep= hint line
        first = f.readline()
        if not first.lower().startswith("sep="):
            f.seek(0)
        reader = csv.reader(f)
        for row in reader:
            if not any(cell.strip() for cell in row):
                continue  # skip blank rows
            if is_section_marker(row):
                if current["rows"]:
                    sections.append(current)
                current = {"name": section_name(row), "rows": []}
            else:
                current["rows"].append(row)

    if current["rows"]:
        sections.append(current)

    return sections


# ---------- Excel writing ----------

def write_overview(ws, rows):
    ws.title = "Overview"
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j == 1:
                cell.font = META_LABEL_FONT
    auto_width(ws)


def write_section(ws, name, rows):
    # Truncate sheet name to 31 chars (Excel limit)
    ws.title = name[:31]

    if not rows:
        return

    # First row: section title banner
    ws.cell(row=1, column=1, value=f"  {name}")
    ncols = max(len(r) for r in rows)
    style_section_row(ws, 1, max(ncols, 1))
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    # Detect if first data row looks like a header (all uppercase or short labels)
    data_start = 2
    header_row = None
    if rows and all(cell == cell.upper() for cell in rows[0] if cell.strip()):
        header_row = rows[0]
        data_rows = rows[1:]
    else:
        data_rows = rows

    row_idx = 2
    if header_row:
        for j, val in enumerate(header_row, start=1):
            ws.cell(row=row_idx, column=j, value=val)
        style_header_row(ws, row_idx, len(header_row))
        row_idx += 1
        data_start = row_idx

    for row in data_rows:
        for j, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=j, value=val)
        row_idx += 1

    auto_width(ws)


def convert(input_path, output_path):
    print(f"Reading:  {input_path}")
    sections = parse_csv(input_path)

    wb = Workbook()
    first = True

    for sec in sections:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()

        if sec["name"] == "Overview":
            write_overview(ws, sec["rows"])
        else:
            write_section(ws, sec["name"], sec["rows"])

    wb.save(output_path)
    print(f"Saved:    {output_path}")
    print(f"Sheets:   {', '.join(ws.title for ws in wb.worksheets)}")


# ---------- Entry point ----------

def find_latest_csv():
    pattern = str(Path.home() / "Downloads" / "weather-station-debug-*.csv")
    files = sorted(glob.glob(pattern))
    if not files:
        return None
    return files[-1]


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        input_csv = sys.argv[1]
    else:
        input_csv = find_latest_csv()
        if not input_csv:
            print("Error: no weather-station-debug-*.csv found in ~/Downloads")
            print("Usage: python3 debug_csv_to_excel.py [input.csv] [output.xlsx]")
            sys.exit(1)

    if len(sys.argv) >= 3:
        output_xlsx = sys.argv[2]
    else:
        output_xlsx = str(Path(input_csv).with_suffix(".xlsx"))

    convert(input_csv, output_xlsx)
